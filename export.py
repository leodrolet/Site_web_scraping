"""export.py — Génération d'un fichier Excel (.xlsx) formaté."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Ordre exact des colonnes du fichier Excel.
COLONNES = [
    "Entreprise", "Prénom", "Nom", "Titre", "Département",
    "Courriel", "Confiance (%)", "Ville", "Province/État",
    "Pays", "Source", "Date de recherche",
]

_BLEU = "1F3A5F"        # en-têtes
_BLEU_PALE = "EAF0F6"   # lignes paires
_BLANC = "FFFFFF"       # lignes impaires


def generer_excel(resultats):
    """
    `resultats` : liste de dicts dont les clés correspondent à COLONNES.
    Retourne les octets d'un fichier .xlsx prêt à être téléchargé.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospection"

    cote = Side(style="thin", color="D0D7DE")
    bordure = Border(left=cote, right=cote, top=cote, bottom=cote)

    # --- En-têtes (gras, fond bleu, texte blanc) ---
    for i, col in enumerate(COLONNES, start=1):
        cellule = ws.cell(row=1, column=i, value=col)
        cellule.font = Font(bold=True, color=_BLANC, size=11)
        cellule.fill = PatternFill("solid", fgColor=_BLEU)
        cellule.alignment = Alignment(horizontal="center", vertical="center")
        cellule.border = bordure

    # --- Données (alternance de couleurs) ---
    for ligne_idx, resultat in enumerate(resultats, start=2):
        couleur = _BLEU_PALE if ligne_idx % 2 == 0 else _BLANC
        for col_idx, col in enumerate(COLONNES, start=1):
            cellule = ws.cell(row=ligne_idx, column=col_idx,
                              value=resultat.get(col, ""))
            cellule.fill = PatternFill("solid", fgColor=couleur)
            cellule.alignment = Alignment(vertical="center")
            cellule.border = bordure

    # --- Largeur des colonnes ajustée au contenu ---
    for col_idx, col in enumerate(COLONNES, start=1):
        largeur_max = len(col)
        for resultat in resultats:
            largeur_max = max(largeur_max, len(str(resultat.get(col, ""))))
        lettre = get_column_letter(col_idx)
        ws.column_dimensions[lettre].width = min(largeur_max + 4, 55)

    ws.freeze_panes = "A2"          # fige la ligne d'en-tête
    ws.row_dimensions[1].height = 22

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
